import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

def ingresar_gastos():
    gastos =[]
    while True:
        fecha = input("Ingresea la fecha del gasto (YYYY-MM-DD) o 'fin' para terminar: ")
        if fecha.lower() == 'fin':
            break
        descripcion = input("Ingrese la descripcion del gasto: ")
        try:
            monto = float(input("Ingrese el monto del gasto: $"))
        except ValueError:
            print("Error: Ingrese un monto valido")
            continue
        gastos.append((fecha, descripcion, monto))
    return gastos

def generar_resumen(gastos):
    if not gastos:
        return None
    
    total_gastos = sum(monto for _, _, monto in gastos)
    gasto_mas_caro = max(gastos, key=lambda x: x[2])
    gasto_mas_barato = min(gastos, key=lambda x: x[2])

    return {
        'Total de gastos': total_gastos,
        'Gasto mas caro': {
            'Fecha': gasto_mas_caro[0],
            'Descripcion': gasto_mas_caro[1],
            'Monto': gasto_mas_caro[2]
        },
        'Gasto mas barato': {
            'Fecha': gasto_mas_barato[0],
            'Descripcion': gasto_mas_barato[1],
            'Monto': gasto_mas_barato[2]
        }
    }

def guardar_informe_en_excel(gastos):
    if not gastos:
        return
    
    libro_excel = openpyxl.Workbook()
    hoja = libro_excel.active
    hoja.title = "Gastos"
    cabeceras = ["Fecha", "Descripción", "Monto"]
    for col_num, header in enumerate(cabeceras, 1):
        col_letra = get_column_letter(col_num)
        celda = hoja[f"{col_letra}1"]
        celda.value = header
        celda.font = Font(bold=True)

    for fila, gasto in enumerate(gastos, 2):
        hoja.cell(row=fila, column=1, value=gasto[0])
        hoja.cell(row=fila, column=2, value=gasto[1])
        hoja.cell(row=fila, column=3, value=gasto[2])

    libro_excel.save("informe_gastos.xlsx")

# Programa principal
if __name__ == "_main_":
    print("Programa de seguimiento de gastos")
    gastos = ingresar_gastos()
    resumen = generar_resumen(gastos)

    if resumen:
        print("\nResumen de gastos:")
        for clave, valor in resumen.items():
            if isinstance(valor, dict):
                print(f"{clave}:")
                for subclave, subvalor in valor.items():
                    print(f"  {subclave}: {subvalor}")
            else:
                print(f"{clave}: {valor}")

        guardar_informe_en_excel(gastos)
        print("\nEl informe de gastos se ha guardado en 'informe_gastos.xlsx'.")
    else:
        print("No se ingresaron gastos.")